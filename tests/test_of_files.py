import csv
import zipfile
from openpyxl import load_workbook
from pypdf import PdfReader
from script_os import ZIP_DIR


def test_xlsx_file():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("HW7xlsx.xlsx") as excel_file:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active
            cell_value = sheet.cell(row=1, column=4).value
            name = 'Последний запуск'
            assert name in cell_value, f"Название колонки: {name} есть в файле"


def test_csv_file():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("HW7csv.csv") as csv_file:
            content = csv_file.read().decode(
                'utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]
            result_list = second_row
            Postback = "ready"
            IDPostback = '15'

            assert result_list[0] == Postback, (f"Название постбэка: {Postback
            } присутствует в таблице {result_list}")
            assert result_list[1] == IDPostback, (f"ID по: {IDPostback
            } присутствует в таблице {IDPostback}")


def test_pdf_file():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("HW7pdf.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[1]
            text = page.extract_text()
            assert 'Тестовый PDF файл' in text